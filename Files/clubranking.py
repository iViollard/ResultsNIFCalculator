import openpyxl
import xlrd
import requests
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors

cellBGReset = PatternFill(start_color='FFFFFF',
                          end_color='FFFFFF',
                          fill_type='solid')
cellClrReset = Font(color='000000')

# loads the workbook - assumes the sussex.xlsx file is in the same directory as the script
wb = openpyxl.load_workbook("sussex.xlsx")

# load first worksheet
ws = wb.worksheets[0]

# downloading the relevant files

# rankings first

# go to the British Fencing Association website and download the rankings file (specified)
url = "https://www.britishfencing.com/wp-content/uploads/2018/10/mf_oct_2018.xls"
downloaded_file = requests.get(url)

# write the contents to a new file called rankings.xls
with open("rankings.xls", 'wb') as file:
    file.write(downloaded_file.content)

# Use xlrd to open older style .xls workbook
rank_wb = xlrd.open_workbook('rankings.xls')

# go to the iViollard website and download the multipliers file (specified)
url = "http://www.cmollard.co.uk/multipliers.xls"
downloaded_file = requests.get(url)

# write the contents to a new file called multipliers.xls
with open("multipliers.xls", 'wb') as file:
    file.write(downloaded_file.content)

# Use xlrd to open older style .xls workbook
multipliers_wb = xlrd.open_workbook('multipliers.xls')

# Get the first sheet of each workbook
rank_ws = rank_wb.sheet_by_index(0)
multipliers_ws = multipliers_wb.sheet_by_index(0)

# Get the total number of rows to be used to create our license list
rows = rank_ws.nrows
rows_lic = multipliers_ws.nrows

# Due to formatting, real numbers don't start until here
startrow = 5

# Create lists of lic numbers
rank_lic = rank_ws.col_values(colx=4, start_rowx=startrow, end_rowx=rows)
mult_lic = multipliers_ws.col_values(colx=0, start_rowx=1, end_rowx=rows_lic)

# print the values in the second column of the first sheet
for row in ws['B1:B{}'.format(ws.max_row)]:
    for cell in row:
        print(cell.value)

# Putting values in same row as "Rank, Name, First name,...." adjust as necessary
ws.cell(2, 7).value = 'Fencer NIF'
ws.cell(2, 8).value = 'Points scored'

nif_total = "=SUM(G3:G{})".format(ws.max_row - 1)
nif_total_list = []


# Define function to lookup NIF and return value
def get_nif(x):
    startrow = 5
    for y in rank_lic:
        if int(x) == y:
            try:
                return int(rank_ws.cell_value(startrow, 9))
            except ValueError:
                pass
        startrow = startrow + 1


for row in ws['D3:D{}'.format(ws.max_row)]:
    for cell in row:
        nif_val = get_nif(cell.value)
        ws.cell(cell.row, 7).value = nif_val
        if nif_val is not None:
            nif_total_list.append(nif_val)


# Define function to lookup multiplier and return value

def get_mult(x):
    startrow = 1
    for y in mult_lic:
        if int(x) == y:
            try:
                return float(multipliers_ws.cell_value(startrow, 1))
            except ValueError:
                pass
        startrow = startrow + 1


for row in ws['A3:A{}'.format(ws.max_row)]:
    for cell in row:
        mult_val = get_mult(cell.value)
        ws.cell(cell.row, 8).value = int(mult_val) * sum(nif_total_list)
        print('mult_val for ' + str(cell.value) + ' is ' + str(mult_val))

# sum of NIF values

Grand_Total_Row = ws.max_row + 1
grand_total = "Grand total"
ws.cell(Grand_Total_Row, 1).value = grand_total
ws.cell(Grand_Total_Row, 7).value = nif_total

# styles - currently setting top row to black text on white bg

ws['A1'].fill = cellBGReset
ws['A1'].font = cellClrReset

# While testing I always save to a different workbook so I don't overwrite my test spreadsheet
wb.save('sussex2.xlsx')